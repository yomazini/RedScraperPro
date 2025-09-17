"""
XLSX Exporter for RedScraperPro
🩸 Export scraped data to Excel format 🩸
"""

import json
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from ..utils.config import Config
from ..utils.logger import Logger


class XLSXExporter:
    """Handles XLSX export functionality"""
    
    def __init__(self, config: Config, logger: Logger):
        self.config = config
        self.logger = logger
        self.output_dir = config.get_output_directory()
    
    def export(self, data: List[Dict[str, Any]], filename: str) -> str:
        """Export data to XLSX format with multiple sheets"""
        if not data:
            raise ValueError("No data to export")
        
        # Ensure filename has .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = self.output_dir / filename
        
        try:
            # Separate posts and comments
            posts = [item for item in data if item.get('type') == 'post']
            comments = [item for item in data if item.get('type') == 'comment']
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets
            if posts:
                self._create_posts_sheet(wb, posts)
            
            if comments:
                self._create_comments_sheet(wb, comments)
            
            # Create summary sheet
            self._create_summary_sheet(wb, data, posts, comments)
            
            # Create statistics sheet
            self._create_statistics_sheet(wb, data)
            
            # Save workbook
            wb.save(filepath)
            
            self.logger.export_complete(str(filepath), len(data))
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"XLSX export failed: {str(e)}")
            raise
    
    def _create_posts_sheet(self, wb: Workbook, posts: List[Dict[str, Any]]):
        """Create posts sheet"""
        ws = wb.create_sheet("Posts")
        
        # Flatten posts data
        flattened_posts = self._flatten_data(posts)
        
        # Convert to DataFrame
        df = pd.DataFrame(flattened_posts)
        
        # Reorder columns for better readability
        priority_columns = [
            'id', 'title', 'author', 'subreddit', 'created_datetime', 
            'score', 'num_comments', 'url', 'selftext', 'permalink'
        ]
        
        # Reorder DataFrame columns
        df = self._reorder_dataframe_columns(df, priority_columns)
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style the sheet
        self._style_worksheet(ws, "Posts Data")
    
    def _create_comments_sheet(self, wb: Workbook, comments: List[Dict[str, Any]]):
        """Create comments sheet"""
        ws = wb.create_sheet("Comments")
        
        # Flatten comments data
        flattened_comments = self._flatten_data(comments)
        
        # Convert to DataFrame
        df = pd.DataFrame(flattened_comments)
        
        # Reorder columns for better readability
        priority_columns = [
            'id', 'author', 'subreddit', 'created_datetime', 
            'score', 'body', 'depth', 'is_reply', 'permalink'
        ]
        
        # Reorder DataFrame columns
        df = self._reorder_dataframe_columns(df, priority_columns)
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style the sheet
        self._style_worksheet(ws, "Comments Data")
    
    def _create_summary_sheet(self, wb: Workbook, data: List[Dict[str, Any]], 
                            posts: List[Dict[str, Any]], comments: List[Dict[str, Any]]):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary", 0)  # Insert as first sheet
        
        # Summary statistics
        summary_data = [
            ["Metric", "Value"],
            ["Total Items", len(data)],
            ["Total Posts", len(posts)],
            ["Total Comments", len(comments)],
            ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            [""],
            ["Score Statistics", ""],
            ["Total Score", sum(item.get('score', 0) for item in data)],
            ["Average Score", round(sum(item.get('score', 0) for item in data) / len(data), 2) if data else 0],
            ["Max Score", max(item.get('score', 0) for item in data) if data else 0],
            ["Min Score", min(item.get('score', 0) for item in data) if data else 0],
        ]
        
        # Add posts statistics
        if posts:
            summary_data.extend([
                [""],
                ["Posts Statistics", ""],
                ["Average Post Score", round(sum(item.get('score', 0) for item in posts) / len(posts), 2)],
                ["Average Comments per Post", round(sum(item.get('num_comments', 0) for item in posts) / len(posts), 2)],
            ])
        
        # Add comments statistics
        if comments:
            summary_data.extend([
                [""],
                ["Comments Statistics", ""],
                ["Average Comment Score", round(sum(item.get('score', 0) for item in comments) / len(comments), 2)],
                ["Average Comment Length", round(sum(len(item.get('body', '')) for item in comments) / len(comments), 2)],
            ])
        
        # Add subreddit breakdown
        subreddit_counts = {}
        for item in data:
            subreddit = item.get('subreddit', 'unknown')
            subreddit_counts[subreddit] = subreddit_counts.get(subreddit, 0) + 1
        
        if subreddit_counts:
            summary_data.extend([
                [""],
                ["Top Subreddits", "Count"],
            ])
            
            # Sort subreddits by count
            sorted_subreddits = sorted(subreddit_counts.items(), key=lambda x: x[1], reverse=True)[:10]
            for subreddit, count in sorted_subreddits:
                summary_data.append([f"r/{subreddit}", count])
        
        # Add data to worksheet
        for row in summary_data:
            ws.append(row)
        
        # Style the summary sheet
        self._style_summary_sheet(ws)
    
    def _create_statistics_sheet(self, wb: Workbook, data: List[Dict[str, Any]]):
        """Create detailed statistics sheet"""
        ws = wb.create_sheet("Statistics")
        
        # Author statistics
        author_stats = {}
        for item in data:
            author = item.get('author', 'unknown')
            if author not in author_stats:
                author_stats[author] = {
                    'posts': 0, 'comments': 0, 'total_score': 0, 
                    'avg_score': 0, 'total_items': 0
                }
            
            if item.get('type') == 'post':
                author_stats[author]['posts'] += 1
            elif item.get('type') == 'comment':
                author_stats[author]['comments'] += 1
            
            author_stats[author]['total_score'] += item.get('score', 0)
            author_stats[author]['total_items'] += 1
        
        # Calculate averages
        for author, stats in author_stats.items():
            if stats['total_items'] > 0:
                stats['avg_score'] = round(stats['total_score'] / stats['total_items'], 2)
        
        # Create author statistics table
        author_data = [
            ["Author", "Posts", "Comments", "Total Items", "Total Score", "Average Score"]
        ]
        
        # Sort authors by total items
        sorted_authors = sorted(author_stats.items(), key=lambda x: x[1]['total_items'], reverse=True)[:20]
        for author, stats in sorted_authors:
            author_data.append([
                author, stats['posts'], stats['comments'], 
                stats['total_items'], stats['total_score'], stats['avg_score']
            ])
        
        # Add author statistics
        ws.append(["Author Statistics (Top 20)"])
        ws.append([])
        for row in author_data:
            ws.append(row)
        
        # Add spacing
        ws.append([])
        ws.append([])
        
        # Subreddit statistics
        subreddit_stats = {}
        for item in data:
            subreddit = item.get('subreddit', 'unknown')
            if subreddit not in subreddit_stats:
                subreddit_stats[subreddit] = {
                    'posts': 0, 'comments': 0, 'total_score': 0, 
                    'avg_score': 0, 'total_items': 0
                }
            
            if item.get('type') == 'post':
                subreddit_stats[subreddit]['posts'] += 1
            elif item.get('type') == 'comment':
                subreddit_stats[subreddit]['comments'] += 1
            
            subreddit_stats[subreddit]['total_score'] += item.get('score', 0)
            subreddit_stats[subreddit]['total_items'] += 1
        
        # Calculate averages
        for subreddit, stats in subreddit_stats.items():
            if stats['total_items'] > 0:
                stats['avg_score'] = round(stats['total_score'] / stats['total_items'], 2)
        
        # Create subreddit statistics table
        subreddit_data = [
            ["Subreddit", "Posts", "Comments", "Total Items", "Total Score", "Average Score"]
        ]
        
        # Sort subreddits by total items
        sorted_subreddits = sorted(subreddit_stats.items(), key=lambda x: x[1]['total_items'], reverse=True)[:20]
        for subreddit, stats in sorted_subreddits:
            subreddit_data.append([
                f"r/{subreddit}", stats['posts'], stats['comments'], 
                stats['total_items'], stats['total_score'], stats['avg_score']
            ])
        
        # Add subreddit statistics
        ws.append(["Subreddit Statistics (Top 20)"])
        ws.append([])
        for row in subreddit_data:
            ws.append(row)
        
        # Style the statistics sheet
        self._style_worksheet(ws, "Detailed Statistics")
    
    def _flatten_data(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Flatten nested dictionaries for Excel export"""
        flattened = []
        
        for item in data:
            flat_item = {}
            self._flatten_dict(item, flat_item)
            flattened.append(flat_item)
        
        return flattened
    
    def _flatten_dict(self, d: Dict[str, Any], parent_key: str = '', sep: str = '_') -> Dict[str, Any]:
        """Recursively flatten a nested dictionary"""
        items = []
        
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                # Handle lists by converting to string representation
                if v and isinstance(v[0], dict):
                    items.append((new_key, json.dumps(v, ensure_ascii=False)))
                else:
                    items.append((new_key, ', '.join(map(str, v)) if v else ''))
            elif v is None:
                items.append((new_key, ''))
            else:
                items.append((new_key, v))
        
        return dict(items)
    
    def _reorder_dataframe_columns(self, df: pd.DataFrame, priority_columns: List[str]) -> pd.DataFrame:
        """Reorder DataFrame columns with priority columns first"""
        existing_priority = [col for col in priority_columns if col in df.columns]
        remaining_columns = [col for col in df.columns if col not in priority_columns]
        
        new_column_order = existing_priority + sorted(remaining_columns)
        return df[new_column_order]
    
    def _style_worksheet(self, ws, title: str):
        """Apply styling to worksheet"""
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # Dark red
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply header styling to first row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze first row
        ws.freeze_panes = "A2"
    
    def _style_summary_sheet(self, ws):
        """Apply special styling to summary sheet"""
        # Title styling
        title_font = Font(bold=True, size=14, color="8B0000")
        section_font = Font(bold=True, color="8B0000")
        
        # Style metric headers and sections
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if "Statistics" in cell.value or cell.value == "Metric":
                        cell.font = section_font
                    elif cell.row == 1:
                        cell.font = title_font
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def export_simple(self, data: List[Dict[str, Any]], filename: str) -> str:
        """Export data to simple XLSX format (single sheet)"""
        if not data:
            raise ValueError("No data to export")
        
        # Ensure filename has .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = self.output_dir / filename
        
        try:
            # Flatten data
            flattened_data = self._flatten_data(data)
            
            # Convert to DataFrame
            df = pd.DataFrame(flattened_data)
            
            # Export to Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Data', index=False)
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Data']
                
                # Style the worksheet
                self._style_worksheet(worksheet, "Data")
            
            self.logger.export_complete(str(filepath), len(data))
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"Simple XLSX export failed: {str(e)}")
            raise
